import requests
import telebot
import gspread
from google.oauth2.service_account import Credentials
import datetime
import pandas as pd
from io import BytesIO
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import time
from threading import Lock
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import csv

# ==================== НАСТРОЙКИ ====================
BOT_TOKEN = os.environ.get('BOT_TOKEN')
SPREADSHEET_NAME = "Посещаемость студентов"
GOOGLE_KEY_FILE = os.path.join(os.path.dirname(__file__), "google_key.json")
GROUP_NAME = "4231133"

# Типы неуважительных пропусков (только они считаются прогулами)
UNRESPECTFUL_STATUSES = ['Отсутствовал']  # ❌

# Количество студентов на одной странице
ITEMS_PER_PAGE = 10
# ===================================================

# ==================== КЛАСС ДЛЯ РАБОТЫ С РАСПИСАНИЕМ ====================
class ScheduleManager:
    """Класс для работы с расписанием из CSV-файла"""
    
    def __init__(self, filename='schedule.csv'):
        self.schedule = {}
        self.filename = filename
        self.load_schedule()
    
    def load_schedule(self):
        """Загружает расписание из CSV-файла"""
        try:
            with open(self.filename, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    day = row['day']
                    week_type = row['week_type']
                    lesson = int(row['lesson'])
                    subgroup = row['subgroup']
                    subject = row['subject']
                    
                    if day not in self.schedule:
                        self.schedule[day] = {'odd': {}, 'even': {}}
                    
                    if lesson not in self.schedule[day][week_type]:
                        self.schedule[day][week_type][lesson] = []
                    
                    self.schedule[day][week_type][lesson].append({
                        'subgroup': subgroup,
                        'subject': subject
                    })
            print(f"✅ Расписание загружено из {self.filename}")
        except FileNotFoundError:
            print(f"❌ Файл {self.filename} не найден")
            print("⚠️ Бот будет работать без расписания")
            self.schedule = {}
        except Exception as e:
            print(f"❌ Ошибка загрузки расписания: {e}")
            self.schedule = {}
    
    def get_week_type(self, date):
        """
        Определяет тип недели:
        - нижняя неделя (even) = нечётные недели (1, 3, 5, ...)
        - верхняя неделя (odd) = чётные недели (2, 4, 6, ...)
        """
        week_num = date.isocalendar()[1]
        
        if week_num % 2 == 1:  # нечётные недели
            return 'even'  # нижняя
        else:               # чётные недели
            return 'odd'   # верхняя
    
    def get_day_lessons(self, date, subgroup='all'):
        """Получает список пар на указанную дату для подгруппы"""
        day_name = date.strftime('%A')  # Monday, Tuesday, etc.
        week_type = self.get_week_type(date)
        
        lessons = []
        if day_name in self.schedule and week_type in self.schedule[day_name]:
            for lesson_num, lesson_data in self.schedule[day_name][week_type].items():
                for item in lesson_data:
                    if item['subgroup'] == 'all' or item['subgroup'] == subgroup:
                        lessons.append({
                            'number': lesson_num,
                            'subject': item['subject'],
                            'for_subgroup': item['subgroup']
                        })
                        break
        return sorted(lessons, key=lambda x: x['number'])
    
    def get_all_lessons_in_month(self, year, month, subgroup='all'):
        """Получает все пары в указанном месяце"""
        start_date = datetime.date(year, month, 1)
        if month == 12:
            end_date = datetime.date(year + 1, 1, 1)
        else:
            end_date = datetime.date(year, month + 1, 1)
        
        current_date = start_date
        all_lessons = []
        
        while current_date < end_date:
            lessons = self.get_day_lessons(current_date, subgroup)
            for lesson in lessons:
                all_lessons.append({
                    'date': current_date,
                    'lesson': lesson['number'],
                    'subject': lesson['subject']
                })
            current_date += datetime.timedelta(days=1)
        
        return all_lessons
    
    def get_lessons_in_range(self, start_date, end_date, subgroup='all'):
        """Получает все пары в указанном диапазоне дат"""
        lessons = []
        current_date = start_date
        
        while current_date <= end_date:
            day_lessons = self.get_day_lessons(current_date, subgroup)
            for lesson in day_lessons:
                lessons.append({
                    'date': current_date,
                    'lesson': lesson['number'],
                    'subject': lesson['subject']
                })
            current_date += datetime.timedelta(days=1)
        
        return lessons
    
    def get_next_unmarked_lesson(self, year, month, marked_lessons, subgroup='all'):
        """Находит следующую неотмеченную пару в указанном месяце"""
        all_lessons = self.get_all_lessons_in_month(year, month, subgroup)
        
        # Сортируем по дате (сначала старые)
        all_lessons.sort(key=lambda x: x['date'])
        
        for lesson in all_lessons:
            date_str = lesson['date'].strftime("%d.%m.%Y")
            lesson_num = lesson['lesson']
            
            # Проверяем, отмечена ли эта пара
            is_marked = any(
                m['date'] == date_str and m['lesson'] == lesson_num 
                for m in marked_lessons
            )
            
            if not is_marked:
                return lesson
        
        return None
# ====================================================

# ==================== НАСТРОЙКА СЕССИИ ====================
session = requests.Session()
retry = Retry(
    total=5,
    read=5,
    connect=5,
    backoff_factor=0.5,
    status_forcelist=(500, 502, 503, 504),
)
adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
session.mount('http://', adapter)
session.mount('https://', adapter)
# ====================================================

# ==================== БЕЗОПАСНОЕ РЕДАКТИРОВАНИЕ СООБЩЕНИЙ ====================
def safe_edit_message(chat_id, message_id, text, reply_markup=None, parse_mode='Markdown'):
    """Безопасное обновление сообщения - игнорирует ошибку 'message is not modified'"""
    try:
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=text,
            parse_mode=parse_mode,
            reply_markup=reply_markup
        )
    except Exception as e:
        if "message is not modified" in str(e).lower():
            pass
        else:
            print(f"⚠️ Ошибка при редактировании: {e}")
# ====================================================

# ==================== БАЗОВОЕ КЭШИРОВАНИЕ ====================
class SheetsCache:
    """Базовый кэш для данных Google Sheets"""
    def __init__(self):
        self.students_cache = []
        self.students_timestamp = 0
        self.attendance_cache = {}
        self.attendance_timestamp = {}
        self.cache_ttl = 30
        self.lock = Lock()
        self.max_retries = 5
        self.base_delay = 1
    
    def _safe_call(self, func, *args, **kwargs):
        for attempt in range(self.max_retries):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                error_str = str(e)
                if '429' in error_str or 'RESOURCE_EXHAUSTED' in error_str:
                    if attempt < self.max_retries - 1:
                        delay = self.base_delay * (2 ** attempt)
                        print(f"⚠️ Превышена квота API. Ожидание {delay} сек... (попытка {attempt + 1}/{self.max_retries})")
                        time.sleep(delay)
                    else:
                        print("❌ Исчерпаны все попытки вызова API")
                        raise e
                else:
                    raise e
    
    def get_students(self):
        with self.lock:
            current_time = time.time()
            if not self.students_cache or current_time - self.students_timestamp > self.cache_ttl:
                try:
                    self.students_cache = self._safe_call(students_sheet.get_all_values)
                    self.students_timestamp = current_time
                    print("📥 Загружен список студентов (кэш обновлён)")
                except Exception as e:
                    if self.students_cache:
                        print("⚠️ Используем устаревший кэш студентов")
                        return self.students_cache
                    raise e
            return self.students_cache
    
    def get_attendance(self, date, lesson):
        key = f"{date}_{lesson}"
        with self.lock:
            current_time = time.time()
            if key not in self.attendance_cache or current_time - self.attendance_timestamp.get(key, 0) > self.cache_ttl:
                try:
                    records = self._safe_call(attendance_sheet.get_all_records)
                    filtered = {}
                    for record in records:
                        if (str(record.get('Дата', '')) == date and
                            str(record.get('Пара', '')) == str(lesson)):
                            student_name = record.get('Студент', '')
                            if student_name:
                                filtered[student_name] = {
                                    'status': record.get('Статус', ''),
                                    'reason': record.get('Причина', '')
                                }
                    self.attendance_cache[key] = filtered
                    self.attendance_timestamp[key] = current_time
                    print(f"📥 Загружены отметки для {date} пара {lesson} (кэш обновлён)")
                except Exception as e:
                    if key in self.attendance_cache:
                        print(f"⚠️ Используем устаревший кэш для {date} пара {lesson}")
                        return self.attendance_cache[key]
                    raise e
            return self.attendance_cache[key]
    
    def clear_attendance_cache(self, date=None, lesson=None):
        with self.lock:
            if date and lesson:
                key = f"{date}_{lesson}"
                self.attendance_cache.pop(key, None)
                self.attendance_timestamp.pop(key, None)
                print(f"🗑️ Очищен кэш для {date} пара {lesson}")
            elif date:
                keys_to_remove = [key for key in self.attendance_cache.keys() if key.startswith(f"{date}_")]
                for key in keys_to_remove:
                    self.attendance_cache.pop(key, None)
                    self.attendance_timestamp.pop(key, None)
                print(f"🗑️ Очищен кэш для всех пар {date}")
            else:
                self.attendance_cache.clear()
                self.attendance_timestamp.clear()
                print("🗑️ Очищен весь кэш отметок")
    
    def clear_students_cache(self):
        with self.lock:
            self.students_cache = []
            self.students_timestamp = 0
            print("🗑️ Очищен кэш студентов")

# ==================== УЛУЧШЕННОЕ КЭШИРОВАНИЕ ====================
class ImprovedSheetsCache(SheetsCache):
    """Улучшенный кэш с принудительным ожиданием между запросами"""
    
    def __init__(self):
        super().__init__()
        self.last_request_time = 0
        self.min_request_interval = 1.1
    
    def _wait_for_rate_limit(self):
        now = time.time()
        time_since_last = now - self.last_request_time
        if time_since_last < self.min_request_interval:
            wait_time = self.min_request_interval - time_since_last
            time.sleep(wait_time)
        self.last_request_time = time.time()
    
    def _safe_call(self, func, *args, **kwargs):
        self._wait_for_rate_limit()
        
        for attempt in range(self.max_retries):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                error_str = str(e)
                if '429' in error_str or 'RESOURCE_EXHAUSTED' in error_str:
                    if attempt < self.max_retries - 1:
                        delay = self.base_delay * (4 ** attempt)
                        print(f"⚠️ Квота API превышена. Ожидание {delay} сек... (попытка {attempt + 1}/{self.max_retries})")
                        time.sleep(delay)
                        self._wait_for_rate_limit()
                    else:
                        print("❌ Исчерпаны все попытки вызова API")
                        raise
                else:
                    raise
# ====================================================

# Расписание пар
LESSON_TIMES = {
    1: "08:00 - 09:30",
    2: "09:40 - 11:10",
    3: "11:50 - 13:20",
    4: "13:30 - 15:00",
    5: "15:40 - 17:10",
    6: "17:20 - 18:50"
}

# Статусы с эмодзи (теперь 'sick' не требует причины)
STATUSES = {
    'present': {'emoji': '✅', 'text': 'Присутствовал'},
    'absent': {'emoji': '❌', 'text': 'Отсутствовал'},
    'sick': {'emoji': '🤒', 'text': 'Болел'},
    'valid': {'emoji': '📄', 'text': 'Уважительная причина'}
}

# Настройка доступа к Google Sheets
scope = ['https://www.googleapis.com/auth/spreadsheets',
         'https://www.googleapis.com/auth/drive']

try:
    from google.oauth2 import service_account
    creds = service_account.Credentials.from_service_account_file(
        GOOGLE_KEY_FILE,
        scopes=scope
    )
    client = gspread.authorize(creds)
    print("✅ Google Таблица подключена!")
except Exception as e:
    print(f"❌ Ошибка подключения к Google: {e}")
    exit()

# Открываем таблицу
try:
    spreadsheet = client.open(SPREADSHEET_NAME)
    attendance_sheet = spreadsheet.worksheet("Посещаемость")
    students_sheet = spreadsheet.worksheet("Студенты")
    print("✅ Google Таблица подключена!")
    
    cache = ImprovedSheetsCache()
    print("✅ Улучшенная система кэширования запущена")
    
    schedule_manager = ScheduleManager('schedule.csv')
    
except Exception as e:
    print(f"❌ Ошибка подключения к Google: {e}")
    exit()

# Создаём бота
bot = telebot.TeleBot(BOT_TOKEN, threaded=True, skip_pending=True)
bot.session = session

# ==================== ХРАНЕНИЕ ТЕКУЩЕГО ВЫБОРА ====================
user_data = {}

def get_user_data(user_id):
    if user_id not in user_data:
        user_data[user_id] = {
            'current_date': datetime.date.today().strftime("%d.%m.%Y"),
            'selected_lessons': set(),
            'selected_subgroup': 'all',
            'marking_mode': False,
            'current_page': 0,
            'students_list': [],
            'selected_students': set()
        }
    return user_data[user_id]

# ==================== ПОЛУЧЕНИЕ ОТМЕЧЕННЫХ ПАР ====================
def get_marked_lessons(year, month):
    """Получает список отмеченных пар за указанный месяц"""
    try:
        records = attendance_sheet.get_all_records()
        marked = []
        seen = set()
        
        # Начало и конец месяца для фильтрации
        start_date = datetime.date(year, month, 1)
        if month == 12:
            end_date = datetime.date(year + 1, 1, 1)
        else:
            end_date = datetime.date(year, month + 1, 1)
        
        for record in records:
            date_str = record.get('Дата', '')
            if not date_str:
                continue
                
            try:
                date = datetime.datetime.strptime(date_str, "%d.%m.%Y").date()
                
                # Проверяем, что дата в нужном месяце
                if start_date <= date < end_date:
                    lesson_num = int(record.get('Пара', 0))
                    pair_key = f"{date_str}_{lesson_num}"
                    
                    if pair_key not in seen:
                        seen.add(pair_key)
                        marked.append({
                            'date': date_str,
                            'lesson': lesson_num
                        })
            except Exception as e:
                print(f"⚠️ Ошибка обработки даты {date_str}: {e}")
                continue
        
        return marked
        
    except Exception as e:
        print(f"❌ Ошибка получения отмеченных пар: {e}")
        return []

# ==================== ГЛАВНОЕ МЕНЮ ====================
@bot.message_handler(commands=['start'])
def start(message):
    user = get_user_data(message.chat.id)
    
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = telebot.types.KeyboardButton('📅 Выбор даты')
    btn2 = telebot.types.KeyboardButton('🔢 Выбрать пары')
    btn3 = telebot.types.KeyboardButton('👥 Подгруппа')
    btn4 = telebot.types.KeyboardButton('📝 Отметить')
    btn5 = telebot.types.KeyboardButton('📊 Состояние')
    btn6 = telebot.types.KeyboardButton('📤 Отчёт')
    markup.add(btn1, btn2, btn3, btn4, btn5, btn6)
    
    subgroup_text = {
        'all': '👥 вся группа',
        '1': '1️⃣ подгруппа 1',
        '2': '2️⃣ подгруппа 2'
    }.get(user['selected_subgroup'], 'не выбрана')
    
    selected_lessons = sorted(user['selected_lessons'])
    lessons_text = f"🔢 *Пары:* {', '.join(map(str, selected_lessons))}" if selected_lessons else "🔢 *Пары:* не выбраны"
    
    # Добавляем информацию о прогрессе
    today = datetime.date.today()
    year = today.year
    month = today.month
    
    all_lessons = schedule_manager.get_all_lessons_in_month(year, month, user['selected_subgroup'])
    marked_lessons = get_marked_lessons(year, month)
    
    total = len(all_lessons)
    marked_count = len(marked_lessons)
    
    month_names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    month_name = month_names[month]
    
    progress_text = f"📊 *Прогресс за {month_name}:* {marked_count} из {total} пар"
    
    bot.send_message(message.chat.id,
                    f"👋 *Система учёта посещаемости*\n"
                    f"👥 *Группа:* {GROUP_NAME}\n"
                    f"👤 *Режим:* {subgroup_text}\n"
                    f"{lessons_text}\n"
                    f"📅 *Дата:* {user['current_date']}\n"
                    f"{progress_text}\n\n"
                    f"Выберите действие:",
                    parse_mode='Markdown',
                    reply_markup=markup)

# ==================== СОСТОЯНИЕ ====================
@bot.message_handler(func=lambda message: message.text == '📊 Состояние')
def show_status(message):
    user = get_user_data(message.chat.id)
    
    today = datetime.date.today()
    year = today.year
    month = today.month
    
    # Получаем все пары в текущем месяце
    all_lessons = schedule_manager.get_all_lessons_in_month(year, month, user['selected_subgroup'])
    
    # Получаем отмеченные пары
    marked_lessons = get_marked_lessons(year, month)
    
    total = len(all_lessons)
    marked_count = len(marked_lessons)
    remaining = total - marked_count
    
    month_names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    month_name = month_names[month]
    
    # Ищем следующую неотмеченную пару (в текущем месяце, даже если в прошлом)
    next_lesson = schedule_manager.get_next_unmarked_lesson(year, month, marked_lessons, user['selected_subgroup'])
    
    status_text = f"📊 *СОСТОЯНИЕ ГРУППЫ*\n\n"
    status_text += f"📅 *{month_name} {year}*\n"
    status_text += f"✅ Отмечено: {marked_count} из {total} пар\n"
    status_text += f"📌 Осталось: {remaining} пар\n\n"
    
    if next_lesson:
        day_name = {
            'Monday': 'пн', 'Tuesday': 'вт', 'Wednesday': 'ср',
            'Thursday': 'чт', 'Friday': 'пт', 'Saturday': 'сб', 'Sunday': 'вс'
        }.get(next_lesson['date'].strftime('%A'), '??')
        
        status_text += f"⏩ *Следующая неотмеченная:*\n"
        status_text += f"📅 {next_lesson['date'].strftime('%d.%m')} ({day_name}) "
        status_text += f"{next_lesson['lesson']} пара - {next_lesson['subject']}\n\n"
        
        markup = telebot.types.InlineKeyboardMarkup()
        markup.add(telebot.types.InlineKeyboardButton(
            "⏩ Перейти к этой паре",
            callback_data=f"goto_{next_lesson['date'].strftime('%d.%m.%Y')}_{next_lesson['lesson']}"
        ))
    else:
        status_text += f"🎉 *Все пары за {month_name} отмечены!*\n\n"
        markup = None
    
    bot.send_message(message.chat.id, status_text, parse_mode='Markdown', reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith('goto_'))
def goto_lesson(call):
    """Переход к указанной паре"""
    user = get_user_data(call.message.chat.id)
    
    # Формат: goto_28.02.2026_3
    parts = call.data.split('_')
    date_str = parts[1]
    lesson_num = int(parts[2])
    
    user['current_date'] = date_str
    user['selected_lessons'] = {lesson_num}
    
    bot.answer_callback_query(call.id, f"✅ Переход к паре {lesson_num} ({date_str})")
    
    # Сразу открываем отметку студентов
    mark_students_for_date(call.message.chat.id, date_str, lesson_num)

def mark_students_for_date(chat_id, date_str, lesson_num):
    """Открывает отметку для конкретной даты и пары"""
    user = get_user_data(chat_id)
    
    try:
        all_students = cache.get_students()
        all_students_list = all_students[1:] if len(all_students) > 1 else []
        
        if user['selected_subgroup'] != 'all':
            students = [s for s in all_students_list 
                       if len(s) >= 3 and str(s[2]) == user['selected_subgroup']]
        else:
            students = all_students_list
        
        if len(students) <= 0:
            bot.send_message(chat_id, "❌ Нет студентов в выбранной подгруппе!")
            return
        
        user['students_list'] = students
        user['selected_students'] = set()
        user['current_page'] = 0
        
        existing_marks = get_existing_marks(date_str, lesson_num)
        user['marking_mode'] = True
        
        show_students_list_with_checkboxes(chat_id, students, existing_marks, 0)
        
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка: {e}")

# ==================== ВЫБОР ДАТЫ ====================
@bot.message_handler(func=lambda message: message.text == '📅 Выбор даты')
def date_choice_menu(message):
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        telebot.types.InlineKeyboardButton("✅ Сегодня", callback_data="date_today"),
        telebot.types.InlineKeyboardButton("📅 Другая дата", callback_data="date_custom")
    )
    
    bot.send_message(message.chat.id,
                    "📅 *Выберите дату:*\n\n"
                    "• ✅ Сегодня — установит текущую дату\n"
                    "• 📅 Другая дата — введите вручную",
                    parse_mode='Markdown',
                    reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == 'date_today')
def set_today(call):
    user = get_user_data(call.message.chat.id)
    user['current_date'] = datetime.date.today().strftime("%d.%m.%Y")
    
    bot.answer_callback_query(call.id, "✅ Дата установлена")
    bot.edit_message_text(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text=f"✅ Установлена сегодняшняя дата: {user['current_date']}",
        parse_mode='Markdown'
    )

@bot.callback_query_handler(func=lambda call: call.data == 'date_custom')
def ask_custom_date(call):
    bot.edit_message_text(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text="📅 *Введите дату*\n\n"
             "Формат: `ДД.ММ.ГГГГ`\n"
             "Пример: `25.02.2026`\n\n"
             "Или нажмите /start для отмены",
        parse_mode='Markdown'
    )
    bot.register_next_step_handler(call.message, process_custom_date)

def process_custom_date(message):
    user = get_user_data(message.chat.id)
    try:
        datetime.datetime.strptime(message.text, "%d.%m.%Y")
        user['current_date'] = message.text
        bot.send_message(message.chat.id, f"✅ Дата установлена: {message.text}")
    except ValueError:
        bot.send_message(message.chat.id, "❌ Неверный формат! Используйте ДД.ММ.ГГГГ")

# ==================== ВЫБОР ПАР ====================
@bot.message_handler(func=lambda message: message.text == '🔢 Выбрать пары')
def choose_lessons(message):
    user = get_user_data(message.chat.id)
    
    try:
        current_date = datetime.datetime.strptime(user['current_date'], "%d.%m.%Y").date()
    except:
        current_date = datetime.date.today()
    
    available_lessons = schedule_manager.get_day_lessons(
        current_date, 
        user['selected_subgroup']
    )
    
    if not available_lessons:
        bot.send_message(message.chat.id,
                        "❌ На выбранную дату нет пар в расписании")
        return
    
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    
    for lesson in available_lessons:
        lesson_num = lesson['number']
        subject = lesson['subject']
        
        if lesson_num in user.get('selected_lessons', set()):
            btn_text = f"✅ {lesson_num} - {subject}"
        else:
            btn_text = f"{lesson_num} - {subject}"
        
        markup.add(
            telebot.types.InlineKeyboardButton(
                btn_text,
                callback_data=f"toggle_lesson_{lesson_num}"
            )
        )
    
    markup.add(
        telebot.types.InlineKeyboardButton("✅ Выбрать все", callback_data="lessons_all"),
        telebot.types.InlineKeyboardButton("❌ Очистить все", callback_data="lessons_clear")
    )
    
    markup.add(
        telebot.types.InlineKeyboardButton("📌 Готово", callback_data="lessons_done")
    )
    
    selected = user.get('selected_lessons', set())
    selected_text = f"✅ *Выбрано пар:* {len(selected)}" if selected else "❌ *Ничего не выбрано*"
    
    schedule_text = "\n".join([f"{l['number']}. {l['subject']}" for l in available_lessons])
    
    bot.send_message(message.chat.id,
                    f"🔢 *ВЫБОР ПАР*\n\n"
                    f"{selected_text}\n\n"
                    f"*Расписание на {user['current_date']}:*\n{schedule_text}\n\n"
                    f"*Нажимайте на пары, чтобы выбрать/снять выбор*",
                    parse_mode='Markdown',
                    reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith('toggle_lesson_'))
def toggle_lesson(call):
    user = get_user_data(call.message.chat.id)
    lesson_num = int(call.data.split('_')[2])
    
    if 'selected_lessons' not in user:
        user['selected_lessons'] = set()
    
    if lesson_num in user['selected_lessons']:
        user['selected_lessons'].remove(lesson_num)
        bot.answer_callback_query(call.id, f"❌ Пара {lesson_num} снята")
    else:
        user['selected_lessons'].add(lesson_num)
        bot.answer_callback_query(call.id, f"✅ Пара {lesson_num} выбрана")
    
    update_lessons_display(call)

def update_lessons_display(call):
    user = get_user_data(call.message.chat.id)
    
    try:
        current_date = datetime.datetime.strptime(user['current_date'], "%d.%m.%Y").date()
    except:
        current_date = datetime.date.today()
    
    available_lessons = schedule_manager.get_day_lessons(
        current_date, 
        user['selected_subgroup']
    )
    
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    
    for lesson in available_lessons:
        lesson_num = lesson['number']
        subject = lesson['subject']
        
        if lesson_num in user['selected_lessons']:
            btn_text = f"✅ {lesson_num} - {subject}"
        else:
            btn_text = f"{lesson_num} - {subject}"
        
        markup.add(
            telebot.types.InlineKeyboardButton(
                btn_text,
                callback_data=f"toggle_lesson_{lesson_num}"
            )
        )
    
    markup.add(
        telebot.types.InlineKeyboardButton("✅ Выбрать все", callback_data="lessons_all"),
        telebot.types.InlineKeyboardButton("❌ Очистить все", callback_data="lessons_clear")
    )
    
    markup.add(
        telebot.types.InlineKeyboardButton("📌 Готово", callback_data="lessons_done")
    )
    
    selected = user['selected_lessons']
    selected_text = f"✅ *Выбрано пар:* {len(selected)}" if selected else "❌ *Ничего не выбрано*"
    
    schedule_text = "\n".join([f"{l['number']}. {l['subject']}" for l in available_lessons])
    
    safe_edit_message(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text=f"🔢 *ВЫБОР ПАР*\n\n"
             f"{selected_text}\n\n"
             f"*Расписание на {user['current_date']}:*\n{schedule_text}\n\n"
             f"*Нажимайте на пары, чтобы выбрать/снять выбор*",
        parse_mode='Markdown',
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == 'lessons_all')
def lessons_all(call):
    user = get_user_data(call.message.chat.id)
    
    try:
        current_date = datetime.datetime.strptime(user['current_date'], "%d.%m.%Y").date()
    except:
        current_date = datetime.date.today()
    
    available_lessons = schedule_manager.get_day_lessons(
        current_date, 
        user['selected_subgroup']
    )
    
    user['selected_lessons'] = {l['number'] for l in available_lessons}
    bot.answer_callback_query(call.id, f"✅ Выбраны все пары ({len(available_lessons)})")
    
    update_lessons_display(call)

@bot.callback_query_handler(func=lambda call: call.data == 'lessons_clear')
def lessons_clear(call):
    user = get_user_data(call.message.chat.id)
    user['selected_lessons'] = set()
    bot.answer_callback_query(call.id, "❌ Выбор очищен")
    update_lessons_display(call)

@bot.callback_query_handler(func=lambda call: call.data == 'lessons_done')
def lessons_done(call):
    user = get_user_data(call.message.chat.id)
    
    if not user.get('selected_lessons'):
        bot.answer_callback_query(call.id, "❌ Выберите хотя бы одну пару!")
        return
    
    selected = sorted(user['selected_lessons'])
    selected_text = ", ".join(map(str, selected))
    
    bot.answer_callback_query(call.id, f"✅ Выбраны пары: {selected_text}")
    
    safe_edit_message(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text=f"✅ *Настройки установлены*\n\n"
             f"📅 *Дата:* {user['current_date']}\n"
             f"🔢 *Выбранные пары:* {selected_text}\n\n"
             f"Теперь можно *отметить студентов* 👇",
        parse_mode='Markdown'
    )

# ==================== ВЫБОР ПОДГРУППЫ ====================
@bot.message_handler(func=lambda message: message.text == '👥 Подгруппа')
def choose_subgroup(message):
    user = get_user_data(message.chat.id)
    
    markup = telebot.types.InlineKeyboardMarkup(row_width=3)
    markup.add(
        telebot.types.InlineKeyboardButton(
            "👥 Вся группа",
            callback_data="subgroup_all"
        ),
        telebot.types.InlineKeyboardButton(
            "1️⃣ Подгруппа 1",
            callback_data="subgroup_1"
        ),
        telebot.types.InlineKeyboardButton(
            "2️⃣ Подгруппа 2",
            callback_data="subgroup_2"
        )
    )
    
    current = {
        'all': '👥 Вся группа',
        '1': '1️⃣ Подгруппа 1',
        '2': '2️⃣ Подгруппа 2'
    }.get(user['selected_subgroup'], 'не выбрана')
    
    bot.send_message(message.chat.id,
                    f"👥 *Выбор подгруппы*\n\n"
                    f"Текущий выбор: {current}\n\n"
                    f"Выберите, кого хотите отмечать:",
                    parse_mode='Markdown',
                    reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith('subgroup_'))
def set_subgroup(call):
    user = get_user_data(call.message.chat.id)
    subgroup = call.data.split('_')[1]
    user['selected_subgroup'] = subgroup
    
    subgroup_text = {
        'all': 'вся группа',
        '1': 'подгруппа 1',
        '2': 'подгруппа 2'
    }.get(subgroup, 'не выбрана')
    
    bot.answer_callback_query(call.id, f"✅ Выбрана {subgroup_text}")
    bot.edit_message_text(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text=f"✅ *Подгруппа выбрана*\n\n"
             f"Теперь вы будете отмечать: {subgroup_text}",
        parse_mode='Markdown'
    )

# ==================== ОТМЕТКА СТУДЕНТОВ ====================
@bot.message_handler(func=lambda message: message.text == '📝 Отметить')
def mark_students(message):
    user = get_user_data(message.chat.id)
    
    if not user.get('selected_lessons'):
        bot.send_message(message.chat.id, 
                        "❌ *Сначала выберите пары!*\n"
                        "Нажмите 🔢 Выбрать пары",
                        parse_mode='Markdown')
        return
    
    try:
        all_students = cache.get_students()
        all_students_list = all_students[1:] if len(all_students) > 1 else []
        
        if user['selected_subgroup'] != 'all':
            students = [s for s in all_students_list 
                       if len(s) >= 3 and str(s[2]) == user['selected_subgroup']]
        else:
            students = all_students_list
        
        if len(students) <= 0:
            bot.send_message(message.chat.id, "❌ Нет студентов в выбранной подгруппе!")
            return
        
        user['students_list'] = students
        user['selected_students'] = set()
        user['current_page'] = 0
        
        existing_marks = {}
        for lesson in user['selected_lessons']:
            marks = get_existing_marks(user['current_date'], lesson)
            for student, data in marks.items():
                if student not in existing_marks:
                    existing_marks[student] = data
        
        user['marking_mode'] = True
        
        selected_lessons = sorted(user['selected_lessons'])
        lessons_text = ", ".join(map(str, selected_lessons))
        
        subgroup_text = {
            'all': 'вся группа',
            '1': 'подгруппа 1',
            '2': 'подгруппа 2'
        }.get(user['selected_subgroup'], 'не выбрана')
        
        bot.send_message(message.chat.id,
                        f"📌 *Отметка*\n"
                        f"👥 {subgroup_text}\n"
                        f"🔢 *Пары:* {lessons_text}\n"
                        f"📅 *Дата:* {user['current_date']}\n\n"
                        f"*Отметки будут применены ко ВСЕМ выбранным парам!*",
                        parse_mode='Markdown')
        
        show_students_list_with_checkboxes(message.chat.id, students, existing_marks, 0)
        
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {e}")

# ==================== ПОЛУЧЕНИЕ СУЩЕСТВУЮЩИХ ОТМЕТОК ====================
def get_existing_marks(date, lesson):
    try:
        return cache.get_attendance(date, lesson)
    except Exception as e:
        print(f"❌ Ошибка получения отметок: {e}")
        return {}

# ==================== СОХРАНЕНИЕ ЗАПИСИ ====================
def save_attendance_record(date, lessons, student, status, reason, force_overwrite=True):
    """Сохраняет запись о посещении для одной или нескольких пар
    Если force_overwrite=True, удаляет старые записи перед сохранением"""
    try:
        if isinstance(lessons, (list, set)):
            lesson_list = list(lessons)
        else:
            lesson_list = [lessons]
        
        time.sleep(1.1)
        records = attendance_sheet.get_all_values()
        
        rows_to_delete = []
        rows_to_add = []
        
        for lesson in lesson_list:
            # Всегда удаляем старые записи для этого студента на эту дату и пару
            for i, row in enumerate(records):
                if (i > 0 and len(row) >= 4 and
                    str(row[0]) == date and
                    str(row[1]) == str(lesson) and
                    str(row[3]) == student):
                    rows_to_delete.append(i + 1)
            
            time_now = datetime.datetime.now().strftime("%H:%M")
            rows_to_add.append([
                date,
                lesson,
                GROUP_NAME,
                student,
                status,
                reason,
                time_now
            ])
        
        if rows_to_delete:
            for row_num in sorted(rows_to_delete, reverse=True):
                attendance_sheet.delete_rows(row_num)
            print(f"🗑️ Удалено {len(rows_to_delete)} записей")
        
        if rows_to_add:
            for row in rows_to_add:
                attendance_sheet.append_row(row)
            print(f"📝 Добавлено {len(rows_to_add)} записей")
        
        for lesson in lesson_list:
            cache.clear_attendance_cache(date, lesson)
        
        return len(rows_to_add)
    except Exception as e:
        print(f"❌ Ошибка сохранения: {e}")
        return 0

# ==================== ПРИМЕНЕНИЕ БОЛЬНИЧНОГО НА ПЕРИОД ====================
def apply_sick_leave(user, student_name, start_date, end_date):
    """Применяет статус 'Болел' ко всем парам в указанном диапазоне,
    перезаписывая любые предыдущие отметки"""
    lessons_in_range = schedule_manager.get_lessons_in_range(
        start_date, end_date, user['selected_subgroup']
    )
    
    updated_count = 0
    for lesson in lessons_in_range:
        # Сохраняем с force_overwrite=True, чтобы перезаписать старые отметки
        save_attendance_record(
            lesson['date'].strftime("%d.%m.%Y"),
            [lesson['lesson']],
            student_name,
            'Болел',
            '-',
            force_overwrite=True
        )
        updated_count += 1
    
    return updated_count

@bot.callback_query_handler(func=lambda call: call.data == 'sick_leave')
def sick_leave_period(call):
    user = get_user_data(call.message.chat.id)
    
    if not user.get('selected_students'):
        bot.answer_callback_query(call.id, "❌ Сначала выберите студентов")
        return
    
    msg = bot.send_message(
        call.message.chat.id,
        f"📅 *Введите период больничного*\n\n"
        f"Формат: `ДД.ММ.ГГГГ-ДД.ММ.ГГГГ`\n"
        f"Пример: `01.03.2026-10.03.2026`\n\n"
        f"👥 Будет применено для {len(user['selected_students'])} студентов\n"
        f"📊 Система автоматически перезапишет все отметки в этом периоде на 'Болел'"
    )
    bot.register_next_step_handler(msg, process_sick_leave)

def process_sick_leave(message):
    user = get_user_data(message.chat.id)
    
    try:
        # Парсим период
        date_str = message.text.strip().split('-')
        if len(date_str) != 2:
            raise ValueError("Неверный формат")
            
        start_date = datetime.datetime.strptime(date_str[0].strip(), "%d.%m.%Y").date()
        end_date = datetime.datetime.strptime(date_str[1].strip(), "%d.%m.%Y").date()
        
        if end_date < start_date:
            bot.send_message(message.chat.id, "❌ Конечная дата раньше начальной!")
            return
        
        total_updated = 0
        for idx in user['selected_students']:
            student_name = get_student_by_index(user, idx)
            if student_name:
                updated = apply_sick_leave(user, student_name, start_date, end_date)
                total_updated += updated
        
        # Очищаем кэш отметок
        cache.clear_attendance_cache()
        
        # Формируем сообщение о результате
        day_count = (end_date - start_date).days + 1
        lessons_count = total_updated // len(user['selected_students']) if user['selected_students'] else 0
        
        bot.send_message(
            message.chat.id,
            f"✅ *Больничный применён*\n\n"
            f"👥 *Студентов:* {len(user['selected_students'])}\n"
            f"📅 *Период:* {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n"
            f"📆 *Дней в периоде:* {day_count}\n"
            f"📊 *Всего обновлено отметок:* {total_updated}\n"
            f"📌 *Пар на студента:* {lessons_count}"
        )
        
        # Очищаем выбор
        user['selected_students'] = set()
        
        # Предлагаем перейти к следующей неотмеченной
        offer_next_unmarked(message.chat.id, user)
        
    except ValueError as e:
        bot.send_message(message.chat.id, 
                        "❌ Неверный формат! Используйте: `ДД.ММ.ГГГГ-ДД.ММ.ГГГГ`\n"
                        "Пример: `01.03.2026-10.03.2026`")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {e}")

# ==================== ПРЕДЛОЖЕНИЕ СЛЕДУЮЩЕЙ ПАРЫ ====================
def offer_next_unmarked(chat_id, user):
    """Предлагает перейти к следующей неотмеченной паре"""
    today = datetime.date.today()
    year = today.year
    month = today.month
    
    marked_lessons = get_marked_lessons(year, month)
    next_lesson = schedule_manager.get_next_unmarked_lesson(
        year, month, marked_lessons, user['selected_subgroup']
    )
    
    if next_lesson:
        day_name = {
            'Monday': 'пн', 'Tuesday': 'вт', 'Wednesday': 'ср',
            'Thursday': 'чт', 'Friday': 'пт', 'Saturday': 'сб', 'Sunday': 'вс'
        }.get(next_lesson['date'].strftime('%A'), '??')
        
        markup = telebot.types.InlineKeyboardMarkup()
        markup.add(
            telebot.types.InlineKeyboardButton(
                "✅ Да, перейти",
                callback_data=f"goto_{next_lesson['date'].strftime('%d.%m.%Y')}_{next_lesson['lesson']}"
            ),
            telebot.types.InlineKeyboardButton(
                "❌ Нет, позже",
                callback_data="cancel_next"
            )
        )
        
        bot.send_message(
            chat_id,
            f"📅 *Следующая неотмеченная пара:*\n"
            f"{next_lesson['date'].strftime('%d.%m')} ({day_name}) "
            f"{next_lesson['lesson']} пара - {next_lesson['subject']}\n\n"
            f"Перейти к отметке?",
            reply_markup=markup
        )

@bot.callback_query_handler(func=lambda call: call.data == 'cancel_next')
def cancel_next(call):
    bot.answer_callback_query(call.id, "❌ Отменено")
    bot.delete_message(call.message.chat.id, call.message.message_id)

# ==================== СОЗДАНИЕ КЛАВИАТУРЫ СТУДЕНТОВ ====================
def create_students_markup(students, existing_marks, page, selected_students):
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    
    selected_count = len(selected_students)
    
    if selected_count > 0:
        # Первая строка: ✅ и ❌
        markup.add(
            telebot.types.InlineKeyboardButton(
                f"✅ Присутствовал",
                callback_data="quick_present"
            ),
            telebot.types.InlineKeyboardButton(
                f"❌ Отсутствовал",
                callback_data="quick_absent"
            )
        )
        
        # Вторая строка: 🤒 и 📄
        markup.add(
            telebot.types.InlineKeyboardButton(
                f"🤒 Болел",
                callback_data="quick_sick"
            ),
            telebot.types.InlineKeyboardButton(
                f"📄 Уважительная",
                callback_data="quick_valid"
            )
        )
        
        # Третья строка: Больничный на период
        markup.add(
            telebot.types.InlineKeyboardButton(
                f"📅 Больничный на период",
                callback_data="sick_leave"
            )
        )
    
    total_students = len(students)
    total_pages = (total_students + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    
    if page < 0:
        page = 0
    elif page >= total_pages:
        page = total_pages - 1
    
    start = page * ITEMS_PER_PAGE
    end = min(start + ITEMS_PER_PAGE, total_students)
    
    for idx_in_list in range(start, end):
        student = students[idx_in_list]
        if len(student) >= 2:
            student_name = student[1]
            
            if student_name in existing_marks:
                status_info = existing_marks[student_name]
                status_text = status_info['status']
                status_emoji = '❓'
                for code, info in STATUSES.items():
                    if info['text'] == status_text:
                        status_emoji = info['emoji']
                        break
                if status_info.get('reason') and status_info['reason'] != '-':
                    status_emoji = f"{status_emoji}📝"
            else:
                status_emoji = '⬜'
            
            checkbox = "☑️" if idx_in_list in selected_students else "◻️"
            
            display_name = student_name
            if len(display_name) > 12:
                display_name = display_name[:12] + "…"
            
            markup.add(
                telebot.types.InlineKeyboardButton(
                    f"{checkbox} {status_emoji} {display_name}",
                    callback_data=f"toggle_{idx_in_list}"
                )
            )
    
    nav_buttons = []
    if page > 0:
        nav_buttons.append(telebot.types.InlineKeyboardButton("◀ Предыдущая", callback_data="page_prev"))
    if page < total_pages - 1:
        nav_buttons.append(telebot.types.InlineKeyboardButton("Следующая ▶", callback_data="page_next"))
    if nav_buttons:
        markup.add(*nav_buttons)
    
    markup.add(
        telebot.types.InlineKeyboardButton("❌ Снять все выборы", callback_data="clear_selection"),
        telebot.types.InlineKeyboardButton("🔄 Обновить", callback_data="refresh_list")
    )
    
    markup.add(
        telebot.types.InlineKeyboardButton("💾 СОХРАНИТЬ И ВЫЙТИ", callback_data="save_exit")
    )
    
    return markup

# ==================== ОТМЕТКА СТУДЕНТОВ С ЧЕКБОКСАМИ ====================
def show_students_list_with_checkboxes(chat_id, students, existing_marks, page=None):
    user = get_user_data(chat_id)
    
    if 'selected_students' not in user:
        user['selected_students'] = set()
    
    if page is None:
        page = user.get('current_page', 0)
    else:
        user['current_page'] = page
    
    total_students = len(students)
    total_pages = (total_students + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    
    if total_pages == 0:
        page = 0
    elif page < 0:
        page = 0
    elif page >= total_pages:
        page = total_pages - 1
    user['current_page'] = page
    
    markup = create_students_markup(students, existing_marks, page, user['selected_students'])
    
    selected_count = len(user['selected_students'])
    selected_text = f"✅ *Выбрано:* {selected_count} студентов\n" if selected_count > 0 else ""
    
    lessons_text = ""
    if user.get('selected_lessons'):
        selected_lessons = sorted(user['selected_lessons'])
        lessons_text = f"🔢 *Пары:* {', '.join(map(str, selected_lessons))}\n"
    
    page_info = f"📄 Страница {page+1} из {total_pages}" if total_pages > 0 else "📄 Нет студентов"
    
    # Добавляем прогресс за день
    day_progress = f"📊 Прогресс за день: {len(selected_lessons)} из {len(selected_lessons)} пар\n" if selected_lessons else ""
    
    bot.send_message(
        chat_id,
        f"📝 *ОТМЕТКА ПОСЕЩАЕМОСТИ*\n\n"
        f"👥 *Группа:* {GROUP_NAME}\n"
        f"📅 *Дата:* {user['current_date']}\n"
        f"{lessons_text}"
        f"{day_progress}"
        f"{selected_text}"
        f"{page_info}\n\n"
        f"*Как отмечать:*\n"
        f"1. Нажмите на студента, чтобы выбрать ☑️\n"
        f"2. Выберите статус для ВСЕХ выбранных\n\n"
        f"*Статусы:* ✅ ❌ 🤒 📄\n"
        f"*🤒 Больничный* — можно указать период",
        parse_mode='Markdown',
        reply_markup=markup
    )

# ==================== БЕЗОПАСНОЕ ПОЛУЧЕНИЕ СТУДЕНТА ====================
def get_student_by_index(user, idx):
    if 'students_list' not in user:
        return None
    if idx >= len(user['students_list']):
        return None
    if len(user['students_list'][idx]) < 2:
        return None
    return user['students_list'][idx][1]

# ==================== ОБРАБОТЧИКИ ДЛЯ ОТМЕТКИ ====================
@bot.callback_query_handler(func=lambda call: call.data.startswith('toggle_'))
def toggle_student(call):
    user = get_user_data(call.message.chat.id)
    idx = int(call.data.split('_')[1])
    
    if idx >= len(user.get('students_list', [])):
        bot.answer_callback_query(call.id, "❌ Данные устарели, обновите список")
        refresh_students_list(call.message.chat.id, call.message.message_id)
        return
    
    if idx in user['selected_students']:
        user['selected_students'].remove(idx)
        bot.answer_callback_query(call.id, "❌ Выбор снят")
    else:
        user['selected_students'].add(idx)
        bot.answer_callback_query(call.id, "✅ Студент выбран")
    
    students = user.get('students_list', [])
    existing_marks = {}
    for lesson in user['selected_lessons']:
        marks = get_existing_marks(user['current_date'], lesson)
        for student, data in marks.items():
            if student not in existing_marks:
                existing_marks[student] = data
    
    update_students_message(call.message.chat.id, call.message.message_id, students, existing_marks)

@bot.callback_query_handler(func=lambda call: call.data == 'clear_selection')
def clear_selection(call):
    user = get_user_data(call.message.chat.id)
    user['selected_students'] = set()
    bot.answer_callback_query(call.id, "❌ Все выборы сняты")
    
    students = user.get('students_list', [])
    existing_marks = {}
    for lesson in user['selected_lessons']:
        marks = get_existing_marks(user['current_date'], lesson)
        for student, data in marks.items():
            if student not in existing_marks:
                existing_marks[student] = data
    
    update_students_message(call.message.chat.id, call.message.message_id, students, existing_marks)

def update_students_message(chat_id, message_id, students, existing_marks):
    user = get_user_data(chat_id)
    
    markup = create_students_markup(students, existing_marks, user['current_page'], user['selected_students'])
    selected_count = len(user['selected_students'])
    selected_text = f"✅ *Выбрано:* {selected_count} студентов\n" if selected_count > 0 else ""
    
    lessons_text = ""
    if user.get('selected_lessons'):
        selected_lessons = sorted(user['selected_lessons'])
        lessons_text = f"🔢 *Пары:* {', '.join(map(str, selected_lessons))}\n"
    
    page = user['current_page']
    total_students = len(students)
    total_pages = (total_students + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    page_info = f"📄 Страница {page+1} из {total_pages}" if total_pages > 0 else "📄 Нет студентов"
    
    day_progress = f"📊 Прогресс за день: {len(selected_lessons)} из {len(selected_lessons)} пар\n" if selected_lessons else ""
    
    safe_edit_message(
        chat_id=chat_id,
        message_id=message_id,
        text=f"📝 *ОТМЕТКА ПОСЕЩАЕМОСТИ*\n\n"
             f"👥 *Группа:* {GROUP_NAME}\n"
             f"📅 *Дата:* {user['current_date']}\n"
             f"{lessons_text}"
             f"{day_progress}"
             f"{selected_text}"
             f"{page_info}\n\n"
             f"*Как отмечать:*\n"
             f"1. Нажмите на студента, чтобы выбрать ☑️\n"
             f"2. Выберите статус для ВСЕХ выбранных\n\n"
             f"*Статусы:* ✅ ❌ 🤒 📄\n"
             f"*🤒 Больничный* — можно указать период",
        parse_mode='Markdown',
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data.startswith('quick_'))
def quick_apply_status(call):
    user = get_user_data(call.message.chat.id)
    status_code = call.data.split('_')[1]
    info = STATUSES[status_code]
    
    if not user.get('selected_students'):
        bot.answer_callback_query(call.id, "❌ Нет выбранных студентов")
        return
    
    # Только 'valid' требует причины (уважительная причина)
    if status_code == 'valid':
        user['pending_status'] = {
            'status_code': status_code,
            'status_text': info['text'],
            'students': list(user['selected_students']).copy(),
            'callback_message_id': call.message.message_id
        }
        
        msg = bot.send_message(
            call.message.chat.id,
            f"📝 *Введите причину для {len(user['selected_students'])} студентов:*\n"
            f"Статус: {info['emoji']} {info['text']}\n\n"
            f"Причина будет применена ко всем выбранным студентам."
        )
        bot.register_next_step_handler(msg, save_reason_for_selected)
        return
    
    # Для остальных статусов (present, absent, sick) - без причины
    for idx in user['selected_students']:
        student_name = get_student_by_index(user, idx)
        if student_name:
            save_attendance_record(
                user['current_date'],
                user['selected_lessons'],
                student_name,
                info['text'],
                "-",
                force_overwrite=True
            )
    
    user['selected_students'] = set()
    bot.answer_callback_query(call.id, f"✅ Статус '{info['text']}' применён")
    
    students = user.get('students_list', [])
    existing_marks = {}
    for lesson in user['selected_lessons']:
        marks = get_existing_marks(user['current_date'], lesson)
        for student, data in marks.items():
            if student not in existing_marks:
                existing_marks[student] = data
    
    update_students_message(call.message.chat.id, call.message.message_id, students, existing_marks)
    
    # Предлагаем перейти к следующей неотмеченной
    offer_next_unmarked(call.message.chat.id, user)

def save_reason_for_selected(message):
    user = get_user_data(message.chat.id)
    reason = message.text
    
    if 'pending_status' not in user:
        bot.send_message(message.chat.id, "❌ Ошибка: данные не найдены")
        return
    
    pending = user['pending_status']
    
    for idx in pending['students']:
        student_name = get_student_by_index(user, idx)
        if student_name:
            save_attendance_record(
                user['current_date'],
                user['selected_lessons'],
                student_name,
                pending['status_text'],
                reason,
                force_overwrite=True
            )
    
    user['selected_students'] = set()
    del user['pending_status']
    
    subgroup_text = {
        'all': 'вся группа',
        '1': 'подгруппа 1',
        '2': 'подгруппа 2'
    }.get(user['selected_subgroup'], 'не выбрана')
    
    bot.send_message(
        message.chat.id,
        f"✅ *Отмечено {len(pending['students'])} студентов*\n"
        f"👥 {subgroup_text}\n"
        f"📝 *Причина:* {reason}\n"
        f"🔢 *Пары:* {', '.join(map(str, sorted(user['selected_lessons'])))}"
    )
    
    students = user.get('students_list', [])
    existing_marks = {}
    for lesson in user['selected_lessons']:
        marks = get_existing_marks(user['current_date'], lesson)
        for student, data in marks.items():
            if student not in existing_marks:
                existing_marks[student] = data
    show_students_list_with_checkboxes(message.chat.id, students, existing_marks, user['current_page'])
    
    # Предлагаем перейти к следующей неотмеченной
    offer_next_unmarked(message.chat.id, user)

@bot.callback_query_handler(func=lambda call: call.data == 'back_to_list')
def back_to_list(call):
    refresh_students_list(call.message.chat.id, call.message.message_id)

@bot.callback_query_handler(func=lambda call: call.data == 'refresh_list')
def refresh_list(call):
    refresh_students_list(call.message.chat.id, call.message.message_id)

def refresh_students_list(chat_id, message_id=None):
    user = get_user_data(chat_id)
    
    try:
        all_students = cache.get_students()
        all_students_list = all_students[1:] if len(all_students) > 1 else []
        
        if user['selected_subgroup'] != 'all':
            students = [s for s in all_students_list 
                       if len(s) >= 3 and str(s[2]) == user['selected_subgroup']]
        else:
            students = all_students_list
        
        old_selection = user.get('selected_students', set())
        user['students_list'] = students
        user['selected_students'] = {idx for idx in old_selection if idx < len(students)}
        
        existing_marks = {}
        for lesson in user['selected_lessons']:
            marks = get_existing_marks(user['current_date'], lesson)
            for student, data in marks.items():
                if student not in existing_marks:
                    existing_marks[student] = data
        
        if message_id:
            update_students_message(chat_id, message_id, students, existing_marks)
        else:
            show_students_list_with_checkboxes(chat_id, students, existing_marks, user.get('current_page', 0))
        
    except Exception as e:
        bot.send_message(chat_id, f"❌ Ошибка обновления: {e}")

@bot.callback_query_handler(func=lambda call: call.data == 'save_exit')
def save_and_exit(call):
    user = get_user_data(call.message.chat.id)
    user['marking_mode'] = False
    user['selected_students'] = set()
    
    bot.answer_callback_query(call.id, "✅ Данные сохранены")
    
    selected_lessons = sorted(user['selected_lessons'])
    lessons_text = ", ".join(map(str, selected_lessons)) if selected_lessons else "не выбраны"
    
    safe_edit_message(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text=f"✅ *Данные сохранены!*\n\n"
             f"📅 *Дата:* {user['current_date']}\n"
             f"🔢 *Пары:* {lessons_text}\n"
             f"👥 *Группа:* {GROUP_NAME}\n\n"
             f"Для нового действия нажмите /start",
        parse_mode='Markdown'
    )
    
    # Предлагаем перейти к следующей неотмеченной
    offer_next_unmarked(call.message.chat.id, user)

@bot.callback_query_handler(func=lambda call: call.data == 'page_prev')
def page_prev(call):
    user = get_user_data(call.message.chat.id)
    current_page = user.get('current_page', 0)
    if current_page > 0:
        students = user.get('students_list', [])
        if not students:
            all_students = cache.get_students()
            all_students_list = all_students[1:] if len(all_students) > 1 else []
            
            if user['selected_subgroup'] != 'all':
                students = [s for s in all_students_list 
                           if len(s) >= 3 and str(s[2]) == user['selected_subgroup']]
            else:
                students = all_students_list
            user['students_list'] = students
        
        existing_marks = {}
        for lesson in user['selected_lessons']:
            marks = get_existing_marks(user['current_date'], lesson)
            for student, data in marks.items():
                if student not in existing_marks:
                    existing_marks[student] = data
        
        user['current_page'] = current_page - 1
        update_students_message(call.message.chat.id, call.message.message_id, students, existing_marks)
    else:
        bot.answer_callback_query(call.id, "Вы на первой странице")

@bot.callback_query_handler(func=lambda call: call.data == 'page_next')
def page_next(call):
    user = get_user_data(call.message.chat.id)
    current_page = user.get('current_page', 0)
    students = user.get('students_list', [])
    total_pages = (len(students) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    
    if current_page < total_pages - 1:
        existing_marks = {}
        for lesson in user['selected_lessons']:
            marks = get_existing_marks(user['current_date'], lesson)
            for student, data in marks.items():
                if student not in existing_marks:
                    existing_marks[student] = data
        
        user['current_page'] = current_page + 1
        update_students_message(call.message.chat.id, call.message.message_id, students, existing_marks)
    else:
        bot.answer_callback_query(call.id, "Вы на последней странице")

# ==================== ДОБАВЛЕНИЕ СТУДЕНТА ====================
def save_new_student(message):
    try:
        name = message.text.strip()
        
        if not name:
            bot.send_message(message.chat.id, "❌ Имя не может быть пустым!")
            return
        
        students = students_sheet.get_all_values()
        for student in students[1:]:
            if len(student) >= 2 and student[1] == name:
                bot.send_message(message.chat.id, f"⚠️ Студент '{name}' уже есть в списке!")
                return
        
        students_sheet.append_row([GROUP_NAME, name])
        cache.clear_students_cache()
        
        bot.send_message(message.chat.id,
                        f"✅ *Студент добавлен!*\n\n"
                        f"👤 *{name}*\n"
                        f"👥 *Группа:* {GROUP_NAME}",
                        parse_mode='Markdown')
        
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {e}")

# ==================== ОТЧЁТЫ ====================
@bot.message_handler(func=lambda message: message.text == '📤 Отчёт')
def get_report_menu(message):
    current_month = datetime.date.today().strftime("%m.%Y")
    msg = bot.send_message(message.chat.id,
                          f"📅 *Введите месяц и год для отчёта*\n\n"
                          f"Формат: `ММ.ГГГГ`\n"
                          f"*Пример:* `{current_month}`\n"
                          f"Или введите `текущий` для текущего месяца",
                          parse_mode='Markdown')
    bot.register_next_step_handler(msg, generate_monthly_report)

def generate_monthly_report(message):
    try:
        if message.text.lower() == 'текущий':
            month_year = datetime.date.today().strftime("%m.%Y")
        else:
            month_year = message.text
        
        month, year = map(int, month_year.split('.'))
        
        time.sleep(1.1)
        records = attendance_sheet.get_all_records()
        if not records:
            bot.send_message(message.chat.id, "📭 Нет данных для отчёта")
            return
        
        df = pd.DataFrame(records)
        df['Дата'] = pd.to_datetime(df['Дата'], format='%d.%m.%Y', errors='coerce')
        
        mask = (df['Дата'].dt.month == month) & (df['Дата'].dt.year == year)
        filtered = df[mask]
        
        if filtered.empty:
            bot.send_message(message.chat.id, f"📭 Нет данных за {month_year}")
            return
        
        all_students_data = cache.get_students()
        all_students = [s[1] for s in all_students_data[1:] if len(s) >= 2]
        
        all_dates = sorted(filtered['Дата'].dt.strftime('%d.%m.%Y').unique())
        
        # ЛИСТ ПОСЕЩАЕМОСТИ
        attendance_matrix = []
        for student in all_students:
            row = {'Студент': student}
            student_records = filtered[filtered['Студент'] == student]
            
            for date in all_dates:
                day_records = student_records[student_records['Дата'].dt.strftime('%d.%m.%Y') == date]
                if not day_records.empty:
                    status = day_records.iloc[0]['Статус']
                    if status == 'Присутствовал':
                        row[date] = '✅'
                    elif status == 'Отсутствовал':
                        row[date] = '❌'
                    elif status == 'Болел':
                        row[date] = '🤒'
                    elif status == 'Уважительная причина':
                        row[date] = '📄'
                    else:
                        row[date] = status
                else:
                    row[date] = ''
            attendance_matrix.append(row)
        
        df_attendance = pd.DataFrame(attendance_matrix)
        
        # ЛИСТ СТАТИСТИКИ
        stats_data = []
        for student in all_students:
            student_records = filtered[filtered['Студент'] == student]
            
            total_classes = len(student_records)
            present = len(student_records[student_records['Статус'] == 'Присутствовал'])
            unexcused = len(student_records[student_records['Статус'] == 'Отсутствовал'])
            sick = len(student_records[student_records['Статус'] == 'Болел'])
            excused = len(student_records[student_records['Статус'] == 'Уважительная причина'])
            
            attendance_rate = round(present / total_classes * 100, 1) if total_classes > 0 else 0
            
            stats_data.append({
                'Студент': student,
                'Всего занятий': total_classes,
                '✅ Присутствовал': present,
                '❌ ПРОГУЛЫ': unexcused,
                '🤒 Болел': sick,
                '📄 Уважительная причина': excused,
                '% посещения': attendance_rate
            })
        
        df_stats = pd.DataFrame(stats_data)
        
        # СОЗДАНИЕ EXCEL
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_attendance.to_excel(writer, sheet_name='Посещаемость', index=False)
            df_stats.to_excel(writer, sheet_name='Статистика', index=False)
            
            reasons_df = filtered[filtered['Причина'] != '-']
            if not reasons_df.empty:
                reasons_df = reasons_df[['Дата', 'Пара', 'Студент', 'Статус', 'Причина']]
                reasons_df.to_excel(writer, sheet_name='Причины', index=False)
            
            workbook = writer.book
            worksheet_stats = writer.sheets['Статистика']
            
            # НАСТРОЙКА ШИРИНЫ СТОЛБЦОВ
            worksheet_stats.column_dimensions['A'].width = 25
            worksheet_stats.column_dimensions['B'].width = 15
            worksheet_stats.column_dimensions['C'].width = 18
            worksheet_stats.column_dimensions['D'].width = 15
            worksheet_stats.column_dimensions['E'].width = 12
            worksheet_stats.column_dimensions['F'].width = 20
            worksheet_stats.column_dimensions['G'].width = 15
            
            # ЦВЕТОВАЯ ИНДИКАЦИЯ
            from openpyxl.styles import PatternFill
            
            green_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            for row in range(2, len(df_stats) + 2):
                cell = worksheet_stats.cell(row=row, column=4)
                if cell.value is not None:
                    if cell.value == 0:
                        cell.fill = green_fill
                    elif cell.value <= 10:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
            
            # ЗАГОЛОВКИ
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col in range(1, 8):
                col_letter = get_column_letter(col)
                cell = worksheet_stats[f'{col_letter}1']
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            worksheet_stats.auto_filter.ref = worksheet_stats.dimensions
        
        output.seek(0)
        
        total_unexcused = df_stats['❌ ПРОГУЛЫ'].sum()
        students_with_absences = len(df_stats[df_stats['❌ ПРОГУЛЫ'] > 0])
        
        caption = (
            f"📊 *ОТЧЁТ ЗА {month_year}*\n\n"
            f"👥 *Группа:* {GROUP_NAME}\n"
            f"📅 *Занятий:* {len(all_dates)}\n"
            f"👤 *Студентов:* {len(all_students)}\n"
            f"❌ *ВСЕГО ПРОГУЛОВ:* {total_unexcused}\n"
            f"⚠️ *Студентов с прогулами:* {students_with_absences}\n\n"
            f"*Цветовая индикация:*\n"
            f"🟢 0 прогулов — без заливки\n"
            f"🟡 ≤ 10 прогулов — жёлтый\n"
            f"🔴 > 10 прогулов — красный"
        )
        
        bot.send_chat_action(message.chat.id, 'upload_document')
        bot.send_document(
            message.chat.id,
            output,
            caption=caption,
            parse_mode='Markdown',
            visible_file_name=f'прогулы_{GROUP_NAME}_{month_year}.xlsx'
        )
        
    except ValueError:
        bot.send_message(message.chat.id, "❌ Неправильный формат! Используйте ММ.ГГГГ")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка генерации отчёта: {str(e)}")

# ==================== ЗАПУСК ====================
if __name__ == "__main__":
    print("=" * 60)
    print("🤖 Бот для учёта посещаемости ЗАПУЩЕН!")
    print("=" * 60)
    print(f"📍 Группа: {GROUP_NAME}")
    print(f"✅ Множественный выбор пар - АКТИВЕН")
    print(f"✅ Поддержка подгрупп - АКТИВНА")
    print(f"✅ Быстрые кнопки статусов")
    print(f"✅ Крупные кнопки для телефона")
    print(f"✅ Объединённый выбор даты")
    print(f"✅ Кнопка 'Состояние' с быстрым переходом")
    print(f"✅ Автопереход к следующей паре")
    print(f"✅ Прогресс при старте")
    print(f"✅ Больничный на период (перезаписывает старые отметки)")
    print(f"✅ Показывает только текущий месяц")
    print(f"✅ Показывает все неотмеченные пары (включая прошедшие)")
    print(f"✅ Расписание с названиями пар")
    print(f"✅ Тип недели: нижняя = нечётные недели")
    print(f"✅ УЛУЧШЕННОЕ КЭШИРОВАНИЕ - АКТИВНО")
    print(f"✅ Батчевые операции - АКТИВНЫ")
    print(f"✅ Автоперезапуск при ошибках - АКТИВЕН")
    print(f"📊 Отчёт: цветовая индикация прогулов")
    print(f"📅 Расписание пар загружено")
    print("=" * 60)
    
    while True:
        try:
            print("🔄 Запуск polling...")
            bot.polling(none_stop=False, interval=1, timeout=30)
        except requests.exceptions.ReadTimeout:
            print("⚠️ Timeout Telegram API, перезапуск через 5 секунд...")
            time.sleep(5)
            continue
        except requests.exceptions.ConnectionError:
            print("⚠️ Ошибка соединения, перезапуск через 10 секунд...")
            time.sleep(10)
            continue
        except Exception as e:
            print(f"❌ Неожиданная ошибка: {e}")
            print("🔄 Перезапуск через 10 секунд...")
            time.sleep(10)
            continue
